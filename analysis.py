import openpyxl
import csv

import pandas as pd
import binascii

# Excelファイルを読み込む
wb = openpyxl.load_workbook('./OUTPUT/SRUM_DUMP_OUTPUT.xlsx')

# シート名のリストを取得
sheet_names = wb.sheetnames


def main():
    """note"""

    # 各シートの内容を取得
    for sheet_name in sheet_names:

        # 実行プログラム
        if sheet_name == "ruDbIdMapTable":
            print(f"[INFO] {sheet_name}")

            # シートデータを出力
            write2csv(sheet_name)

            # Pandasで読み込む
            # CSVファイルを読み込んでDataFrameオブジェクトとして格納
            headers = ["IdType", "IdIndex", "IdBlob"]
            df = pd.read_csv(f"{sheet_name}.csv", names=headers, header=0)

            # 空欄を0にする
            # df.fillna("", inplace=True)

            # 特定の列を1行ずつ出力する
            # df['IdBlob_utf8'] = df['IdBlob']
            for index, row in df.iterrows():
                res = hex_to_str(str(row['IdBlob']))
                df['IdBlob'][index] = res

            df.to_csv(f"{sheet_name}_ana.csv", index=False)

        # CPU/Battery
        if sheet_name == "5C8CF1C7-7257-":
            print(f"[INFO] {sheet_name}")

            # シートデータを出力
            write2csv(sheet_name)

# 16進数表記から文字列に変換する関数
def hex_to_str(hex_str):
    """Converts a hexadecimal string to a UTF-8 string"""

    # 16進数表記であり、かつ16進数表記が偶数桁の場合
    if hex_str and len(hex_str) % 2 == 0:
        print(f"[INFO] {hex_str}")
        # 16進数表記をバイナリ表記に変換
        bin_str = binascii.unhexlify(hex_str)

        # バイナリ表記を文字列に変換
        str_str = bin_str.decode('utf-8')

        return str_str

    return "-"

def write2csv(sheet_name):
    """note"""

    # CSVファイルを開く
    with open(f"{sheet_name}.csv", 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        # シートを選択する
        sheet = wb[sheet_name]

        # 全ての行を取得する
        rows = sheet.iter_rows(values_only=True)

        # 行ごとに処理する
        for row in rows:
            # CSVに書き込む
            writer.writerow(row)

        # シート間に空行を入れる
        writer.writerow([])

    # ファイルを閉じる
    wb.close()


if __name__ == "__main__":

    try:
        main()

    except Exception as err:
        print(f"[ERROR] {err}")
